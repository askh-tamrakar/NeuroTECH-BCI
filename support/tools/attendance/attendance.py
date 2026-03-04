# attendance_gui.py
import tkinter as tk
from tkinter import ttk, messagebox, filedialog, PanedWindow
from datetime import datetime, date
import openpyxl
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
import json
import os
import sys
import calendar
import csv

class AttendanceManager:
    def __init__(self, root):
        self.root = root
        self.root.title("Attendance Manager - NeuroKeys: BCI Typing Project")
        self.root.geometry("1400x900")
        self.root.configure(bg="#f0f0f0")

        # base folder where script lives
        try:
            self.base_dir = os.path.dirname(os.path.abspath(__file__)) or os.getcwd()
        except NameError:
            self.base_dir = os.getcwd()

        self.members = []
        self.attendance = {}  # { member: { "YYYY-MM-DD": "P"/"A"/"L" } }
        self.holidays = []    # list of "YYYY-MM-DD"
        self.current_date = datetime.now().strftime("%Y-%m-%d")
        self.data_file = os.path.join(self.base_dir, "Neurokeys - BCI Typing Project.json")
        self.excel_file = os.path.join(self.base_dir, "Neurokeys - BCI Typing Project.xlsx")
        self.preview_mode = "txt"

        self.load_data()
        self.create_widgets()

    def create_widgets(self):
        # Header Frame
        header_frame = tk.Frame(self.root, bg="#218089", height=60)
        header_frame.pack(fill=tk.X)
        header_frame.pack_propagate(False)

        header_label = tk.Label(
            header_frame,
            text="📋 NeuroKeys: A BCI Tying Project",
            font=("Arial", 18, "bold"),
            bg="#218089",
            fg="white"
        )
        header_label.pack(pady=10)

        # Controls Frame (top)
        controls_frame = tk.Frame(self.root, bg="white")
        controls_frame.pack(fill=tk.X, padx=8, pady=6)

        # Date selection
        date_label = tk.Label(controls_frame, text="📅 Date:", font=("Arial", 10, "bold"), bg="white")
        date_label.grid(row=0, column=0, padx=6, pady=6, sticky="w")
        self.date_entry = tk.Entry(controls_frame, width=15, font=("Arial", 11), relief=tk.SOLID, bd=1)
        self.date_entry.insert(0, self.current_date)
        self.date_entry.grid(row=0, column=1, padx=6, pady=6)
        self.date_entry.bind("<KeyRelease>", lambda e: self.update_date())

        # Add members
        member_label = tk.Label(controls_frame, text="👥 Add Members (comma-separated):", font=("Arial", 10, "bold"), bg="white")
        member_label.grid(row=0, column=2, padx=8, pady=6, sticky="w")
        self.member_input = tk.Entry(controls_frame, width=40, font=("Arial", 10), relief=tk.SOLID, bd=1)
        self.member_input.grid(row=0, column=3, padx=6, pady=6)
        self.member_input.insert(0, "e.g., Member 1, Member 2, Member 3")
        add_btn = tk.Button(controls_frame, text="➕ Add Members", command=self.add_members, bg="#218089", fg="white")
        add_btn.grid(row=0, column=4, padx=6, pady=6)

        # Holidays input
        holiday_label = tk.Label(controls_frame, text="🎉 Holidays (YYYY-MM-DD, comma):", font=("Arial", 10, "bold"), bg="white")
        holiday_label.grid(row=1, column=0, padx=6, pady=6, sticky="w")
        self.holiday_input = tk.Entry(controls_frame, width=60, font=("Arial", 10), relief=tk.SOLID, bd=1)
        self.holiday_input.grid(row=1, column=1, columnspan=3, padx=6, pady=6, sticky="w")
        self.holiday_input.insert(0, ", ".join(self.holidays) if self.holidays else "")
        holiday_btn = tk.Button(controls_frame, text="Set Holidays", command=self.set_holidays, bg="#6b7280", fg="white")
        holiday_btn.grid(row=1, column=4, padx=6, pady=6)

        # Filename + action buttons
        filename_frame = tk.Frame(self.root, bg="white")
        filename_frame.pack(fill=tk.X, padx=8, pady=6)
        filename_label = tk.Label(filename_frame, text="📁 Excel Filename:", font=("Arial", 10, "bold"), bg="white")
        filename_label.pack(side=tk.LEFT, padx=6)
        self.filename_entry = tk.Entry(filename_frame, width=40, font=("Arial", 10), relief=tk.SOLID, bd=1)
        self.filename_entry.pack(side=tk.LEFT, padx=6)
        default_name = os.path.splitext(os.path.basename(self.excel_file))[0]
        self.filename_entry.insert(0, default_name)
        filename_hint = tk.Label(filename_frame, text="(without .xlsx)", bg="white", fg="#666")
        filename_hint.pack(side=tk.LEFT, padx=6)

        # action buttons (Excel update, TXT, CSV export, open)
        btn_frame = tk.Frame(filename_frame, bg="white")
        btn_frame.pack(side=tk.RIGHT)
        excel_btn = tk.Button(btn_frame, text="📊 Update Excel", command=self.update_excel, bg="#22c55e", fg="white")
        excel_btn.pack(side=tk.LEFT, padx=4)
        txt_btn = tk.Button(btn_frame, text="📄 Export TXT", command=self.export_txt, bg="#3b82f6", fg="white")
        txt_btn.pack(side=tk.LEFT, padx=4)
        csv_btn = tk.Button(btn_frame, text="🗂 Export CSV (per-month)", command=self.export_csv_per_month, bg="#10b981", fg="white")
        csv_btn.pack(side=tk.LEFT, padx=4)
        open_btn = tk.Button(btn_frame, text="📂 Open Excel", command=self.open_excel_file, bg="#f59e0b", fg="white")
        open_btn.pack(side=tk.LEFT, padx=4)

        # Stats
        stats_frame = tk.Frame(self.root, bg="white")
        stats_frame.pack(fill=tk.X, padx=8, pady=6)
        self.total_label = tk.Label(stats_frame, text="📊 Total: 0", font=("Arial", 11, "bold"), bg="white")
        self.total_label.pack(side=tk.LEFT, padx=10)
        self.present_label = tk.Label(stats_frame, text="✅ Present: 0", font=("Arial", 11, "bold"), bg="white", fg="#22c55e")
        self.present_label.pack(side=tk.LEFT, padx=10)
        self.absent_label = tk.Label(stats_frame, text="❌ Absent: 0", font=("Arial", 11, "bold"), bg="white", fg="#ef4444")
        self.absent_label.pack(side=tk.LEFT, padx=10)
        self.percentage_label = tk.Label(stats_frame, text="📈 Attendance: 0%", font=("Arial", 11, "bold"), bg="white")
        self.percentage_label.pack(side=tk.LEFT, padx=10)

        # Main panes: left = attendance list; right = preview
        main_paned = PanedWindow(self.root, orient=tk.HORIZONTAL, sashrelief=tk.RAISED)
        main_paned.pack(fill=tk.BOTH, expand=True, padx=8, pady=6)

        # LEFT: attendance canvas
        left_frame = tk.Frame(main_paned, bg="white")
        main_paned.add(left_frame, minsize=500)

        self.canvas_frame = tk.Frame(left_frame, bg="white", relief=tk.SUNKEN, bd=1)
        self.canvas_frame.pack(fill=tk.BOTH, expand=True, padx=6, pady=6)

        self.canvas = tk.Canvas(self.canvas_frame, bg="white", highlightthickness=0)
        scrollbar = ttk.Scrollbar(self.canvas_frame, orient=tk.VERTICAL, command=self.canvas.yview)
        self.scrollable_frame = tk.Frame(self.canvas, bg="white")

        self.scrollable_frame.bind("<Configure>", lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all")))
        self.canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        self.canvas.configure(yscrollcommand=scrollbar.set)
        self.canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.attendance_frame = self.scrollable_frame

        # RIGHT: preview / controls stacked vertically
        right_frame = tk.Frame(main_paned, bg="#f0f0f0")
        main_paned.add(right_frame, minsize=500)

        # Top of right: preview mode tabs buttons horizontally then content below (vertical)
        tab_buttons_frame = tk.Frame(right_frame, bg="#f0f0f0")
        tab_buttons_frame.pack(fill=tk.X, pady=(6, 0), padx=6)
        self.txt_tab_btn = tk.Button(tab_buttons_frame, text="📝 TXT Preview", command=lambda: self.show_preview("txt"), bg="#218089", fg="white")
        self.txt_tab_btn.pack(side=tk.LEFT, padx=4)
        self.excel_tab_btn = tk.Button(tab_buttons_frame, text="📊 Excel Preview", command=lambda: self.show_preview("excel"), bg="#f0f0f0")
        self.excel_tab_btn.pack(side=tk.LEFT, padx=4)

        # Preview area (stacked vertically) — REPLACEMENT: adds scrollbars and monospace font
        preview_area = tk.Frame(right_frame, bg="#f9f9f9", relief=tk.SUNKEN, bd=1)
        preview_area.pack(fill=tk.BOTH, expand=True, padx=6, pady=6)

        # horizontal + vertical scrollbars
        preview_vscroll = ttk.Scrollbar(preview_area, orient=tk.VERTICAL)
        preview_hscroll = ttk.Scrollbar(preview_area, orient=tk.HORIZONTAL)

        # Text widget for preview — monospace, no wrap so columns align, connect scrollbars
        self.preview_text = tk.Text(preview_area,
                                    font=("Consolas", 16),
                                    bg="#ffffff",
                                    wrap=tk.NONE,
                                    yscrollcommand=preview_vscroll.set,
                                    xscrollcommand=preview_hscroll.set)
        preview_vscroll.config(command=self.preview_text.yview)
        preview_hscroll.config(command=self.preview_text.xview)

        # layout: text fills and scrollbars attached (use grid, NOT pack)
        self.preview_text.grid(row=0, column=0, sticky="nsew")
        preview_vscroll.grid(row=0, column=1, sticky="ns")
        preview_hscroll.grid(row=1, column=0, sticky="ew")

        # configure grid weights so preview_text expands
        preview_area.grid_rowconfigure(0, weight=1)
        preview_area.grid_columnconfigure(0, weight=1)

        # allow mouse wheel when cursor is over the preview_text
        def _bind_preview_mousewheel(widget):
            widget.bind("<Enter>", lambda e: widget.bind_all("<MouseWheel>", self._on_preview_mousewheel))
            widget.bind("<Leave>", lambda e: widget.unbind_all("<MouseWheel>"))

        _bind_preview_mousewheel(self.preview_text)

        # Improved mouse wheel handling (widget-scoped, cross-platform)
        def _on_canvas_mousewheel(event):
            # Windows/macOS: event.delta, Linux: event.num
            if hasattr(event, "delta") and event.delta:
                lines = int(-1 * (event.delta / 120))
                # keep min movement of 1
                if lines == 0:
                    lines = -1 if event.delta > 0 else 1
                self.canvas.yview_scroll(lines, "units")
            else:
                if getattr(event, "num", None) == 5:
                    self.canvas.yview_scroll(1, "units")
                elif getattr(event, "num", None) == 4:
                    self.canvas.yview_scroll(-1, "units")

        def _on_preview_mousewheel_event(event):
            if hasattr(event, "delta") and event.delta:
                lines = int(-1 * (event.delta / 120))
                if lines == 0:
                    lines = -1 if event.delta > 0 else 1
                self.preview_text.yview_scroll(lines, "units")
            else:
                if getattr(event, "num", None) == 5:
                    self.preview_text.yview_scroll(1, "units")
                elif getattr(event, "num", None) == 4:
                    self.preview_text.yview_scroll(-1, "units")

        # Scoped bind helper: on Enter bind wheel events, on Leave unbind them
        def _bind_widget_wheel(widget, handler):
            widget.bind("<Enter>", lambda e: (
                widget.bind_all("<MouseWheel>", handler),
                widget.bind_all("<Button-4>", handler),
                widget.bind_all("<Button-5>", handler)
            ))
            widget.bind("<Leave>", lambda e: (
                widget.unbind_all("<MouseWheel>"),
                widget.unbind_all("<Button-4>"),
                widget.unbind_all("<Button-5>")
            ))

        _bind_widget_wheel(self.canvas, _on_canvas_mousewheel)
        _bind_widget_wheel(self.preview_text, _on_preview_mousewheel_event)

        # Render initial
        self.render_attendance()
        self.update_preview()

    # ---- Holidays ----
    def set_holidays(self):
        text = self.holiday_input.get().strip()
        if not text:
            self.holidays = []
        else:
            items = [s.strip() for s in text.split(",") if s.strip()]
            valid = []
            for it in items:
                try:
                    datetime.strptime(it, "%Y-%m-%d")
                    valid.append(it)
                except Exception:
                    messagebox.showwarning("Invalid date", f"Skipped invalid holiday: {it}")
            self.holidays = valid
        self.save_data()
        self.update_preview()
        messagebox.showinfo("Holidays set", f"{len(self.holidays)} holidays configured")

    # ---- Rendering / marking UI (checkbox removed) ----
    def render_attendance(self):
        for widget in self.attendance_frame.winfo_children():
            widget.destroy()

        if not self.members:
            empty_label = tk.Label(self.attendance_frame, text="👥 Add members to get started", font=("Arial", 12), bg="white", fg="#999")
            empty_label.pack(pady=30)
            return

        for member in self.members:
            # current status
            status_text = self.attendance.get(member, {}).get(self.current_date, "A")

            member_frame = tk.Frame(self.attendance_frame, bg="#fafafa", relief=tk.FLAT, bd=1, highlightthickness=1, highlightbackground="#e0e0e0")
            member_frame.pack(fill=tk.X, pady=4, padx=5)

            label = tk.Label(member_frame, text=member, font=("Arial", 11, "bold"), bg="#fff7e0", fg="#333", anchor="w")
            label.pack(side=tk.LEFT, padx=8, fill=tk.X, expand=True)

            status_color = "#0000FF" if status_text == "P" else ("#FF8800" if status_text == "L" else "#FF0000")
            status_bg = "#e8f5e9" if status_text == "P" else ("#fff4e0" if status_text == "L" else "#ffebee")

            status_label = tk.Label(member_frame, text=status_text, font=("Arial", 11, "bold"), bg=status_bg, fg=status_color, width=4, relief=tk.RAISED, bd=1, cursor="hand2")
            status_label.pack(side=tk.RIGHT, padx=12, pady=6)

            # clicking cycles A -> P -> L -> A
            status_label.bind("<Button-1>", lambda e, m=member: self.cycle_status(m))

    def cycle_status(self, member):
        if member not in self.attendance:
            self.attendance[member] = {}
        cur = self.attendance[member].get(self.current_date, "A")
        order = ["A", "P", "L"]
        next_status = order[(order.index(cur) + 1) % len(order)]
        self.attendance[member][self.current_date] = next_status
        self.save_data()
        self.render_attendance()
        self.update_preview()
        self.update_stats()

    def add_members(self):
        input_text = self.member_input.get().strip()
        if not input_text or input_text == "e.g., Member 1, Member 2, Member 3":
            messagebox.showwarning("⚠️ Warning", "Please enter member names")
            return

        new_members = [m.strip() for m in input_text.split(",") if m.strip()]
        added = [m for m in new_members if m not in self.members]
        if not added:
            messagebox.showinfo("ℹ️ Info", "All members already exist")
            return

        self.members.extend(added)
        self.members.sort()
        for member in added:
            if member not in self.attendance:
                self.attendance[member] = {}
        self.save_data()
        self.render_attendance()
        self.member_input.delete(0, tk.END)
        self.update_stats()
        self.update_preview()
        messagebox.showinfo("✅ Success", f"Added {len(added)} member(s)")

    # ---- preview / stats / txt preview ----
    def update_date(self):
        try:
            self.current_date = self.date_entry.get()
            datetime.strptime(self.current_date, "%Y-%m-%d")
            self.attendance_title_text = f"✔️ Mark Attendance for {self.current_date}"
            self.render_attendance()
            self.update_preview()
            self.update_stats()
        except Exception:
            pass

    def update_stats(self):
        total = len(self.members)
        present = sum(1 for m in self.members if self.attendance.get(m, {}).get(self.current_date) == "P")
        absent = sum(1 for m in self.members if self.attendance.get(m, {}).get(self.current_date, "A") == "A")
        leave = sum(1 for m in self.members if self.attendance.get(m, {}).get(self.current_date) == "L")
        percentage = int((present / total * 100)) if total > 0 else 0
        self.total_label.config(text=f"📊 Total: {total}")
        self.present_label.config(text=f"✅ Present: {present}")
        self.absent_label.config(text=f"❌ Absent: {absent}")
        self.percentage_label.config(text=f"📈 Attendance: {percentage}%")

    def show_preview(self, mode: str):
        """
        Switch preview tab between 'txt' and 'excel', update button styles and refresh preview.
        """
        # set mode
        self.preview_mode = mode

        # update tab button appearance if buttons exist
        try:
            if mode == "txt":
                self.txt_tab_btn.config(bg="#218089", fg="white", relief=tk.RAISED, bd=2)
                self.excel_tab_btn.config(bg="#f0f0f0", fg="#333", relief=tk.RAISED, bd=1)
            else:
                self.txt_tab_btn.config(bg="#f0f0f0", fg="#333", relief=tk.RAISED, bd=1)
                self.excel_tab_btn.config(bg="#218089", fg="white", relief=tk.RAISED, bd=2)
        except Exception:
            # safe: if UI widgets not present yet, ignore styling
            pass

        # refresh preview contents
        try:
            self.update_preview()
        except Exception as e:
            # fallback: show error in preview box so you can debug
            try:
                self.preview_text.config(state=tk.NORMAL)
                self.preview_text.delete(1.0, tk.END)
                self.preview_text.insert(1.0, f"Error updating preview: {e}")
                self.preview_text.config(state=tk.DISABLED)
            except Exception:
                # last resort: print to console
                print("Error in show_preview -> update_preview:", e)


    def update_preview(self):
        self.preview_text.config(state=tk.NORMAL)
        self.preview_text.delete(1.0, tk.END)
        if self.preview_mode == "txt":
            self.show_txt_preview()
        else:
            self.show_excel_preview()
        self.preview_text.config(state=tk.DISABLED)

    def show_txt_preview(self):
        present_members = [m for m in self.members if self.attendance.get(m, {}).get(self.current_date) == "P"]
        absent_members = [m for m in self.members if self.attendance.get(m, {}).get(self.current_date) != "P"]
        preview = f"Date: {self.current_date} : Total members: {len(self.members)}\n"
        preview += f"Present ({len(present_members)}) : Absent ({len(absent_members)})\n"
        preview += "\nMember                          Status\n" + "─" * 40 + "\n"
        for member in self.members:
            status = self.attendance.get(member, {}).get(self.current_date, "A")
            preview += f"{member:<30} {status}\n"
        self.preview_text.insert(1.0, preview)

    def show_excel_preview(self):
        """Robust, untruncated preview that formats columns and shows all rows with scroll support."""
        filename = self.filename_entry.get().strip() or os.path.splitext(os.path.basename(self.excel_file))[0]
        excel_file = os.path.join(self.base_dir, f"{filename}.xlsx")

        # Clear preview area first
        try:
            self.preview_text.config(state=tk.NORMAL)
            self.preview_text.delete(1.0, tk.END)
        except Exception:
            pass

        if not os.path.exists(excel_file):
            preview = f"Excel file will be created at: {excel_file}\n\n"
            preview += "Format Preview:\n"
            preview += "S.n. | Student name | NOVEMBER (27..30 as subcolumns) | DECEMBER | ...\n"
            preview += "─" * 120 + "\n"
            preview += " 1   | Akshansh     |  P | A | P | ...\n"
            preview += "...\n"
            self.preview_text.insert(1.0, preview)
            self.preview_text.config(state=tk.DISABLED)
            return

        try:
            wb = openpyxl.load_workbook(excel_file, data_only=True)
            ws = wb.active

            # safe getter for merged cells (returns top-left if empty)
            def get_cell_val(r, c):
                cell = ws.cell(row=r, column=c)
                if cell.value is not None:
                    return cell.value
                # merged ranges
                for m in ws.merged_cells.ranges:
                    if m.min_row <= r <= m.max_row and m.min_col <= c <= m.max_col:
                        tl = ws.cell(row=m.min_row, column=m.min_col)
                        return tl.value
                return ""

            max_col = ws.max_column or 1
            max_row = ws.max_row or 1

            # build a 2-row header (row1 + row2) plus data rows from row3..max_row
            header1 = [str(get_cell_val(1, c) or "") for c in range(1, max_col + 1)]
            header2 = [str(get_cell_val(2, c) or "") for c in range(1, max_col + 1)]

            # collect rows
            data_rows = []
            for r in range(3, max_row + 1):
                row = [str(ws.cell(row=r, column=c).value or "") for c in range(1, max_col + 1)]
                data_rows.append(row)

            # compute column widths (cap width to avoid extreme lengths)
            cols = [header1, header2] + data_rows
            col_widths = []
            for c in range(max_col):
                maxw = 0
                for row in cols:
                    if c < len(row):
                        maxw = max(maxw, len(str(row[c])))
                maxw = min(maxw, 25)  # cap
                col_widths.append(max(3, maxw + 2))

            # helper to format a row
            def fmt_row(row_vals):
                pieces = []
                for i, val in enumerate(row_vals):
                    w = col_widths[i]
                    pieces.append(str(val)[:w - 2].ljust(w))
                return "| " + " | ".join(pieces) + " |"

            # compose preview lines
            lines = []
            lines.append(f"📊 Real-time Excel Preview: {excel_file}")
            lines.append("=" * (sum(col_widths) + 4 * len(col_widths)))
            lines.append(fmt_row(header1))
            lines.append(fmt_row(header2))
            lines.append("-" * (sum(col_widths) + 4 * len(col_widths)))

            # add all data rows (no truncation)
            for r in data_rows:
                lines.append(fmt_row(r))

            # if nothing in data rows
            if not data_rows:
                lines.append("(No member rows yet — click Update Excel to create file)")

            self.preview_text.insert(1.0, "\n".join(lines))
            self.preview_text.config(state=tk.DISABLED)
            wb.close()
        except Exception as exc:
            self.preview_text.config(state=tk.NORMAL)
            self.preview_text.delete(1.0, tk.END)
            self.preview_text.insert(1.0, f"Error reading Excel file for preview:\n{str(exc)}\n\nClick 'Update Excel' to (re)create the workbook.")
            self.preview_text.config(state=tk.DISABLED)

    # ---- Excel update (keeps holidays marked as H) ----
    def update_excel(self):
        if not self.members:
            messagebox.showwarning("⚠️ Warning", "Add members first")
            return
        try:
            filename = self.filename_entry.get().strip() or os.path.splitext(os.path.basename(self.excel_file))[0]
            self.excel_file = os.path.join(self.base_dir, f"{filename}.xlsx")
            date_obj = datetime.strptime(self.current_date, "%Y-%m-%d")
            year = date_obj.year
            month_num = date_obj.month
            month_name = date_obj.strftime("%B").upper()
            day = int(date_obj.day)
            # start day special-case for November
            start_day = 27 if month_name == "NOVEMBER" else 1
            days_in_month = calendar.monthrange(year, month_num)[1]
            day_numbers = list(range(start_day, days_in_month + 1))
            num_day_cols = len(day_numbers)

            if os.path.exists(self.excel_file):
                wb = openpyxl.load_workbook(self.excel_file)
                ws = wb.active
            else:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Attendance"
                ws['A1'] = "S.n."
                ws['B1'] = "Student name"
                header_fill = PatternFill(start_color="9AD3DB", end_color="9AD3DB", fill_type="solid")
                for cell in ['A1', 'B1']:
                    ws[cell].fill = header_fill
                    ws[cell].font = Font(bold=True, color="000000")
                    ws[cell].alignment = Alignment(horizontal="center", vertical="center")

            # find or add month block
            month_start = None
            month_len = None
            for col in range(3, ws.max_column + 1):
                val = ws.cell(row=1, column=col).value
                if val and isinstance(val, str) and val.strip().upper() == month_name:
                    month_start = col
                    length = 0
                    for c in range(col, ws.max_column + 1):
                        r2 = ws.cell(row=2, column=c).value
                        if r2 and (str(r2).isdigit()):
                            length += 1
                        else:
                            break
                    month_len = length if length > 0 else num_day_cols
                    break

            if month_start is None:
                month_start = ws.max_column + 1
                month_len = num_day_cols
                for i, d in enumerate(day_numbers):
                    ws.cell(row=2, column=month_start + i, value=str(d))
                ws.merge_cells(start_row=1, start_column=month_start, end_row=1, end_column=month_start + num_day_cols - 1)
                ws.cell(row=1, column=month_start, value=month_name)
                ws.cell(row=1, column=month_start).fill = PatternFill(start_color="C77AAE", end_color="C77AAE", fill_type="solid")
                ws.cell(row=1, column=month_start).font = Font(bold=True, color="FFFFFF")
                ws.cell(row=1, column=month_start).alignment = Alignment(horizontal="center", vertical="center")
                for c in range(month_start, month_start + num_day_cols):
                    ws.cell(row=2, column=c).fill = PatternFill(start_color="F2E6F2", end_color="F2E6F2", fill_type="solid")
                    ws.cell(row=2, column=c).font = Font(bold=True, color="000000")
                    ws.cell(row=2, column=c).alignment = Alignment(horizontal="center", vertical="center")
            else:
                # ensure month covers needed days (insert if necessary)
                existing_day_numbers = []
                for c in range(month_start, month_start + month_len):
                    val = ws.cell(row=2, column=c).value
                    if val and str(val).isdigit():
                        existing_day_numbers.append(int(val))
                existing_min = min(existing_day_numbers) if existing_day_numbers else None
                existing_max = max(existing_day_numbers) if existing_day_numbers else None

                # if missing earlier days -> insert columns
                if existing_min is None or existing_min > start_day:
                    needed_left = (existing_min - start_day) if existing_min else (start_day - 1)
                    if needed_left > 0:
                        ws.insert_cols(month_start, amount=needed_left)
                        for i in range(needed_left):
                            dval = start_day + i
                            ws.cell(row=2, column=month_start + i, value=str(dval))
                            ws.cell(row=2, column=month_start + i).fill = PatternFill(start_color="F2E6F2", end_color="F2E6F2", fill_type="solid")
                            ws.cell(row=2, column=month_start + i).font = Font(bold=True, color="000000")
                            ws.cell(row=2, column=month_start + i).alignment = Alignment(horizontal="center", vertical="center")
                        try:
                            ws.unmerge_cells(start_row=1, start_column=month_start, end_row=1, end_column=month_start + month_len - 1)
                        except Exception:
                            pass
                        new_span = month_len + needed_left
                        ws.merge_cells(start_row=1, start_column=month_start, end_row=1, end_column=month_start + new_span - 1)
                        ws.cell(row=1, column=month_start).value = month_name
                        ws.cell(row=1, column=month_start).fill = PatternFill(start_color="C77AAE", end_color="C77AAE", fill_type="solid")
                        ws.cell(row=1, column=month_start).font = Font(bold=True, color="FFFFFF")
                        ws.cell(row=1, column=month_start).alignment = Alignment(horizontal="center", vertical="center")
                        month_len = new_span

                # if missing later days -> insert columns at the end of block
                if existing_max is None or existing_max < days_in_month:
                    missing = days_in_month - (existing_max or 0)
                    if missing > 0:
                        insert_at = month_start + month_len
                        ws.insert_cols(insert_at, amount=missing)
                        for i in range(missing):
                            dval = (existing_max or 0) + i + 1
                            ws.cell(row=2, column=insert_at + i, value=str(dval))
                            ws.cell(row=2, column=insert_at + i).fill = PatternFill(start_color="F2E6F2", end_color="F2E6F2", fill_type="solid")
                            ws.cell(row=2, column=insert_at + i).font = Font(bold=True, color="000000")
                            ws.cell(row=2, column=insert_at + i).alignment = Alignment(horizontal="center", vertical="center")
                        try:
                            ws.unmerge_cells(start_row=1, start_column=month_start, end_row=1, end_column=month_start + month_len - 1)
                        except Exception:
                            pass
                        ws.merge_cells(start_row=1, start_column=month_start, end_row=1, end_column=month_start + month_len + missing - 1)
                        ws.cell(row=1, column=month_start).value = month_name
                        ws.cell(row=1, column=month_start).fill = PatternFill(start_color="C77AAE", end_color="C77AAE", fill_type="solid")
                        ws.cell(row=1, column=month_start).font = Font(bold=True, color="FFFFFF")
                        ws.cell(row=1, column=month_start).alignment = Alignment(horizontal="center", vertical="center")
                        month_len = month_len + missing

                for c in range(month_start, month_start + month_len):
                    ws.cell(row=2, column=c).fill = PatternFill(start_color="F2E6F2", end_color="F2E6F2", fill_type="solid")
                    ws.cell(row=2, column=c).font = Font(bold=True, color="000000")
                    ws.cell(row=2, column=c).alignment = Alignment(horizontal="center", vertical="center")

            # day_col index (column corresponding to the selected current_date day)
            day_col = None
            for c in range(month_start, month_start + (num_day_cols if 'num_day_cols' in locals() else month_len)):
                val = ws.cell(row=2, column=c).value
                if val and str(val).isdigit() and int(val) == day:
                    day_col = c
                    break
            if day_col is None:
                # fallback: map by offset from start_day
                day_col = month_start + (day - start_day)

            # style A1/B1
            header_fill = PatternFill(start_color="9AD3DB", end_color="9AD3DB", fill_type="solid")
            ws['A1'].fill = header_fill
            ws['B1'].fill = header_fill
            ws['A1'].font = Font(bold=True, color="000000")
            ws['B1'].font = Font(bold=True, color="000000")
            ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
            ws['B1'].alignment = Alignment(horizontal="center", vertical="center")

            # populate member rows & style (write status into the column for current_date only)
            for idx, member in enumerate(self.members, start=1):
                found_row = None
                for r in range(3, ws.max_row + 1):
                    val = ws.cell(row=r, column=2).value
                    if val and str(val).strip() == member:
                        found_row = r
                        break
                if not found_row:
                    found_row = ws.max_row + 1 if ws.max_row >= 3 else 3
                    ws.cell(row=found_row, column=1, value=idx)
                    ws.cell(row=found_row, column=2, value=member)
                else:
                    ws.cell(row=found_row, column=1, value=idx)

                status = self.attendance.get(member, {}).get(self.current_date, "A")

                # write status only into the day_col (i.e. the column for the selected current_date)
                # (Other columns are left alone so month history remains)
                try:
                    day_val = ws.cell(row=2, column=day_col).value
                    cell_day = int(day_val)
                    cell_date_iso = date(year, month_num, cell_day).isoformat()  # 'YYYY-MM-DD'
                except Exception:
                    cell_date_iso = None

                cell = ws.cell(row=found_row, column=day_col)
                # If this column date is configured as a holiday -> write 'H' and grey it
                if cell_date_iso and (cell_date_iso in self.holidays):
                    cell.value = "H"
                    cell.font = Font(color="000000", bold=True)
                    cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
                else:
                    # Normal status write (P/A/L)
                    cell.value = status
                    if status == "P":
                        cell.font = Font(color="0000FF", bold=True)   # blue
                    elif status == "A":
                        cell.font = Font(color="FF0000", bold=True)   # red
                    elif status == "L":
                        cell.font = Font(color="FF8800", bold=True)   # orange
                    else:
                        cell.font = Font(color="000000")
                    # green background for date cells (non-holiday)
                    cell.fill = PatternFill(start_color="E6F9E6", end_color="E6F9E6", fill_type="solid")

                # cream background for member name cell
                ws.cell(row=found_row, column=2).fill = PatternFill(start_color="FFF7E0", end_color="FFF7E0", fill_type="solid")
                ws.cell(row=found_row, column=2).alignment = Alignment(horizontal="left", vertical="center")

            # IMPORTANT: ensure every holiday column across the block is written as "H" for all member rows
            # (previously only the selected day column was marked). This loop will set 'H' into all holiday columns.
            for c in range(month_start, month_start + (num_day_cols if 'num_day_cols' in locals() else month_len)):
                col_day_val = ws.cell(row=2, column=c).value
                try:
                    col_day = int(col_day_val)
                    col_date_iso = date(year, month_num, col_day).isoformat()
                except Exception:
                    col_date_iso = None

                if col_date_iso and (col_date_iso in self.holidays):
                    # mark entire column cells (row 3 .. max_row) as holiday 'H'
                    for r in range(3, ws.max_row + 1):
                        # don't overwrite student name column or serial column
                        # write H into date cell (col c)
                        ws.cell(row=r, column=c).value = "H"
                        ws.cell(row=r, column=c).font = Font(color="000000", bold=True)
                        ws.cell(row=r, column=c).fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")

            # ensure all date columns in block are green (except holidays)
            for r in range(3, ws.max_row + 1):
                for c in range(month_start, month_start + (num_day_cols if 'num_day_cols' in locals() else month_len)):
                    col_day_val = ws.cell(row=2, column=c).value
                    try:
                        col_day = int(col_day_val)
                        col_date = date(year, month_num, col_day).isoformat()
                    except Exception:
                        col_date = None
                    if col_date and col_date in self.holidays:
                        # already set to grey above; keep grey
                        ws.cell(row=r, column=c).fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
                    else:
                        if ws.cell(row=r, column=c).value is None:
                            ws.cell(row=r, column=c).value = ""
                        ws.cell(row=r, column=c).fill = PatternFill(start_color="E6F9E6", end_color="E6F9E6", fill_type="solid")

            # borders & dims
            max_col = ws.max_column
            for r in range(1, ws.max_row + 1):
                for c in range(1, max_col + 1):
                    cell = ws.cell(row=r, column=c)
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    if r >= 3:
                        cell.border = Border(left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'),
                                             top=Side(style='thin', color='CCCCCC'), bottom=Side(style='thin', color='CCCCCC'))
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 30
            for col_idx in range(3, max_col + 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 5
            ws.row_dimensions[1].height = 26
            ws.row_dimensions[2].height = 22

            wb.save(self.excel_file)
            self.update_preview()
            messagebox.showinfo("✅ Success", f"Excel file updated: {self.excel_file}")
        except Exception as e:
            messagebox.showerror("❌ Error", f"Failed to update Excel:\n{e}")

    # ---- Export CSV per month ----
    def export_csv_per_month(self):
        """
        Create one CSV file per month present in the workbook.
        CSV columns: S.n., Student name, day1, day2, ...
        Holidays marked with 'H', P/A/L kept as-is.
        """
        try:
            filename = self.filename_entry.get().strip() or os.path.splitext(os.path.basename(self.excel_file))[0]
            excel_file = os.path.join(self.base_dir, f"{filename}.xlsx")
            if not os.path.exists(excel_file):
                messagebox.showwarning("No Excel", "Excel file does not exist. Click Update Excel to generate it first.")
                return

            wb = openpyxl.load_workbook(excel_file, data_only=True)
            ws = wb.active

            # Discover month blocks by scanning row1 headers
            col = 3
            month_blocks = []
            while col <= ws.max_column:
                header = ws.cell(row=1, column=col).value
                if header and isinstance(header, str) and header.strip():
                    # find span (count contiguous numeric day cells in row2 from col)
                    start_col = col
                    length = 0
                    while (col + length) <= ws.max_column and ws.cell(row=2, column=col + length).value and str(ws.cell(row=2, column=col + length).value).isdigit():
                        length += 1
                    if length > 0:
                        month_blocks.append((header.strip(), start_col, length))
                        col = col + length
                        continue
                col += 1

            if not month_blocks:
                messagebox.showinfo("No months", "No month blocks found in Excel.")
                return

            # For each month block, write CSV
            exported = []
            for month_name, start_col, length in month_blocks:
                # create CSV filename: <filename>_<month>.csv
                csv_name = f"{filename}_{month_name}.csv"
                csv_path = os.path.join(self.base_dir, csv_name)
                with open(csv_path, 'w', newline='', encoding='utf-8') as csvf:
                    writer = csv.writer(csvf)
                    # header row: S.n., Student name, <days...>
                    day_headers = [str(ws.cell(row=2, column=start_col + i).value or "") for i in range(length)]
                    writer.writerow(["S.n.", "Student name"] + day_headers)
                    # data rows
                    for r in range(3, ws.max_row + 1):
                        sn = ws.cell(row=r, column=1).value or ""
                        name = ws.cell(row=r, column=2).value or ""
                        rowvals = []
                        for i in range(length):
                            c = start_col + i
                            val = ws.cell(row=r, column=c).value
                            if val is None:
                                val = ""
                            rowvals.append(str(val))
                        writer.writerow([sn, name] + rowvals)
                exported.append(csv_name)

            wb.close()
            messagebox.showinfo("✅ CSV Exported", f"CSV files exported next to script:\n{', '.join(exported)}")
        except Exception as e:
            messagebox.showerror("❌ CSV error", f"Failed to export CSVs:\n{e}")

    # ---- Open excel ----
    def open_excel_file(self):
        filename = self.filename_entry.get().strip() or os.path.splitext(os.path.basename(self.excel_file))[0]
        file_to_open = os.path.join(self.base_dir, f"{filename}.xlsx")
        if not os.path.exists(file_to_open):
            messagebox.showwarning("⚠️ Warning", f"Excel file not found:\n{file_to_open}\n\nClick 'Update Excel' first!")
            return
        try:
            if os.name == 'nt':
                os.startfile(file_to_open)
            else:
                if sys.platform == "darwin":
                    os.system(f'open "{file_to_open}"')
                else:
                    os.system(f'xdg-open "{file_to_open}"')
        except Exception as e:
            messagebox.showerror("❌ Error", f"Failed to open file:\n{e}")

    # ---- Export TXT (unchanged) ----
    def export_txt(self, default_save=False):
        if not self.members:
            messagebox.showwarning("⚠️ Warning", "Add members first")
            return
        present_members = [m for m in self.members if self.attendance.get(m, {}).get(self.current_date) == "P"]
        absent_members = [m for m in self.members if self.attendance.get(m, {}).get(self.current_date) != "P"]
        content_lines = [
            f"Date: {self.current_date} : Total members: {len(self.members)}",
            f"Present ({len(present_members)}): {', '.join(present_members) if present_members else 'None'}",
            f"Absent ({len(absent_members)}): {', '.join(absent_members) if absent_members else 'None'}",
            "",
            "Full column preview:",
            f"{'Member':<30} Status",
            "─" * 40
        ]
        for member in self.members:
            status = self.attendance.get(member, {}).get(self.current_date, "A")
            content_lines.append(f"{member:<30} {status}")
        content = "\n".join(content_lines) + "\n"
        if default_save:
            filename = self.filename_entry.get().strip() or os.path.splitext(os.path.basename(self.excel_file))[0]
            folder = self.base_dir
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            default_name = f"Attendance_{self.current_date}_{timestamp}.txt"
            file_path = os.path.join(folder, default_name)
            try:
                with open(file_path, 'w', encoding='utf-8', newline='\n') as f:
                    f.write(content)
                messagebox.showinfo("✅ Success", f"TXT file auto-saved:\n{file_path}")
            except Exception as e:
                messagebox.showerror("❌ Error", f"Could not auto-save file:\n{e}")
            return
        file_path = filedialog.asksaveasfilename(defaultextension=".txt", filetypes=[("Text files", "*.txt"), ("All files", "*.*")], initialfile=f"Attendance_{self.current_date}.txt", initialdir=self.base_dir)
        if not file_path:
            return
        try:
            with open(file_path, 'w', encoding='utf-8', newline='\n') as f:
                f.write(content)
            messagebox.showinfo("✅ Success", f"TXT file saved!\n\n{file_path}")
        except Exception as e:
            messagebox.showerror("❌ Error", f"Could not save TXT file:\n{e}")

    # ---- Save/load ----
    def save_data(self):
        data = {"members": self.members, "attendance": self.attendance, "holidays": self.holidays}
        try:
            with open(self.data_file, 'w', encoding='utf-8') as f:
                json.dump(data, f, indent=2)
        except Exception:
            pass

    def load_data(self):
        if os.path.exists(self.data_file):
            try:
                with open(self.data_file, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    self.members = data.get("members", [])
                    self.attendance = data.get("attendance", {})
                    self.holidays = data.get("holidays", [])
                    # holiday_input may not exist when called during __init__ prior to create_widgets; that's fine
            except Exception:
                pass

    # legacy canvas mousewheel (kept for compatibility)
    def _on_mousewheel(self, event):
        if event.num == 5 or (hasattr(event, "delta") and event.delta < 0):
            self.canvas.yview_scroll(1, "units")
        elif event.num == 4 or (hasattr(event, "delta") and event.delta > 0):
            self.canvas.yview_scroll(-1, "units")

    def _on_preview_mousewheel(self, event):
        if hasattr(event, "delta") and event.delta < 0:
            self.preview_text.yview_scroll(1, "units")
        elif hasattr(event, "delta") and event.delta > 0:
            self.preview_text.yview_scroll(-1, "units")
        else:
            if getattr(event, "num", None) == 5:
                self.preview_text.yview_scroll(1, "units")
            elif getattr(event, "num", None) == 4:
                self.preview_text.yview_scroll(-1, "units")

if __name__ == "__main__":
    root = tk.Tk()
    app = AttendanceManager(root)
    root.mainloop()
